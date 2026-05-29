#!/usr/bin/env python3
"""
import_to_report.py  –  fold legacy phase outputs into a PCB's unified report.

Imports into results/<part-number>/ (creating/appending the current version):
  • Phase 1 : a test_results_*.csv written by hw_test_runner.write_results_csv
  • Phase 2 : a PowerModuleSweep_*.xlsx written by run_power_module_sweep

Phase-2 verdicts are recomputed switch-aware (bottom = duty complement). The
CSV does not carry per-channel analog readings, so the Analog Readings sheet
stays empty for an imported Phase 1 (only PASS/FAIL + flags are reconstructed).

Usage:
  python import_to_report.py --part-number INVGEN3B1-03 \
      --phase1-csv ../results/test_results_PCB-B1_20260529_1216.csv \
      --pm-xlsx    ../results/PowerModuleSweep_PCB-B1_20260529_1318.xlsx
"""
import os
import re
import csv
import argparse

from openpyxl import load_workbook

from hw_protocol import HwTestProtocol
import run_power_module_sweep as pm
from report_store import update_report

_RESULTS_DIR = os.path.abspath(
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "results"))
SWITCHES = ["UTOP", "UBOT", "VTOP", "VBOT", "WTOP", "WBOT"]


def _num(s):
    """Leading number from a formatted cell ('10.40 ns' → 10.40); NaN if none."""
    if s is None:
        return float("nan")
    m = re.match(r"[-+]?\d*\.?\d+", str(s).strip())
    return float(m.group()) if m else float("nan")


def load_phase1_csv(path):
    proto = HwTestProtocol()
    td = proto.sfHwTestProtocolGetTestsDictionary()
    results = []
    with open(path, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            tk = row["test_key"]
            try:
                code = int(row["test_code"], 16)
            except (ValueError, KeyError):
                code = td.get(tk, (0, ""))[0]
            results.append(dict(
                test_key=tk, test_code=code,
                description=td.get(tk, (code, ""))[1],
                timeout_ms=0, duration_s=float(row.get("duration_s") or 0.0),
                result=row.get("result", ""),
                firmware_result=row.get("firmware_result", ""),
                flags=None, flags_str=row.get("flags_str", "—"),
                echo_received=int(row.get("echo_received") or 0),
                analog_values={}, analog_checks={}, analog_flags={},
                analog_result="N/A", message=row.get("message", ""),
                phase=1, flags_history=[], latched_flags=0, early_exit=False))
    return results


def load_pm_xlsx(path):
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    sweep_str = ""
    records = []
    for row in ws.iter_rows(values_only=True):
        if row and row[0] == "Sweep":
            sweep_str = str(row[1] or "")
        sw = row[0] if row else None
        if sw not in SWITCHES:
            continue
        meas = dict(
            frequency=_num(row[4]) * 1000.0,
            duty=_num(row[5]),
            rise_time=_num(row[6]) * 1e-9,
            fall_time=_num(row[7]) * 1e-9,
            v_high=_num(row[8]), v_low=_num(row[9]),
            v_pp=_num(row[10]), v_pp_avg=_num(row[11]))
        rec = dict(switch=sw, freq_khz=_num(row[1]), duty_pct=_num(row[2]),
                   deadband_ns=_num(row[3]), meas=meas)
        pm.recompute_verdict(rec)          # switch-aware verdict + checks
        records.append(rec)
    return records, sweep_str


def main():
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--part-number", "--pn", required=True)
    p.add_argument("--phase1-csv", default=None)
    p.add_argument("--pm-xlsx", default=None)
    args = p.parse_args()

    results = load_phase1_csv(args.phase1_csv) if args.phase1_csv else None
    pm_records, sweep_str = (load_pm_xlsx(args.pm_xlsx) if args.pm_xlsx
                             else (None, ""))

    meta = {}
    if sweep_str:
        meta["power_module_sweep"] = sweep_str
    bits = []
    if results:
        bits.append("Phase 1 (self)")
    if pm_records:
        bits.append("Phase 2 (power module)")
    if bits:
        meta["sequence"] = " + ".join(bits) + " + Phase 3 (HV)"

    out, version = update_report(
        args.part_number, _RESULTS_DIR, session_meta=meta,
        results=results, pm_records=pm_records)

    print("=" * 72)
    print(f"  IMPORTED INTO {args.part_number}  (V{version})")
    if results is not None:
        npass = sum(1 for r in results if r["result"] == "PASS")
        print(f"  Phase 1 : {len(results)} tests  ({npass} PASS)  ← {os.path.basename(args.phase1_csv)}")
    if pm_records is not None:
        npass = sum(1 for r in pm_records if r["verdict"])
        print(f"  Phase 2 : {len(pm_records)} setpoints  ({npass} PASS)  ← {os.path.basename(args.pm_xlsx)}")
    print(f"  Report  : {out}")
    print("=" * 72)


if __name__ == "__main__":
    main()
