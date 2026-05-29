#!/usr/bin/env python3
"""
run_campaign_multi.py  –  run the HW PCBA test campaign N times in a row
and produce a repeatability comparison.

Usage
-----
  python run_campaign_multi.py                 # 3 runs, default sequence
  python run_campaign_multi.py --runs 5
  python run_campaign_multi.py --sequence A0_A1 --unit-sn PCB-001

For each run it produces  results/test_results_run<N>.csv.
After all runs it writes:
  results/ValidationReport_<SN>_<timestamp>_multi.xlsx
      Sheets:
        Summary             — per-run PASS / FAIL / duration totals
        Repeatability       — per-test PASS pattern across runs
        Run <N>             — full result of each run

Repeatability classifications:
  STABLE PASS   — PASS on every run
  STABLE FAIL   — FAIL on every run
  FLAKY         — different result between runs
"""
import sys
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

import argparse
import os
import datetime
import time
from typing import List, Dict, Any

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config import TEST_SEQUENCE, UNIT_SN, OPERATOR
from hw_test_runner import execute_campaign
from hw_test_criteria import ANLG_LIMITS, channel_name


# ── styles ────────────────────────────────────────────────────────────────
NAVY   = "1F3864"
PASS_F = "E2EFDA"
FAIL_F = "FFE0DC"
WARN_F = "FFF2CC"
STRIPE = "F2F2F2"

THIN  = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def _fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _hdr_row(ws, row_idx, labels, widths=None):
    f = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    fill = _fill(NAVY)
    for ci, lbl in enumerate(labels, 1):
        c = ws.cell(row=row_idx, column=ci, value=lbl)
        c.font, c.fill, c.alignment, c.border = f, fill, CENTER, BORDER
        if widths and ci - 1 < len(widths):
            ws.column_dimensions[get_column_letter(ci)].width = widths[ci - 1]
    ws.row_dimensions[row_idx].height = 22


def _result_fill(result):
    return _fill({"PASS": PASS_F, "FAIL": FAIL_F}.get(result, WARN_F))


def _result_font(result):
    color = {"PASS": "006100", "FAIL": "9C0006"}.get(result, "9C6500")
    return Font(name="Arial", size=10, bold=True, color=color)


# ── multi-run driver ──────────────────────────────────────────────────────

def run_n_campaigns(n_runs: int, sequence: str, unit_sn: str,
                    operator: str, results_dir: str,
                    inter_run_pause_s: float = 2.0):
    """Run the campaign n_runs times. Returns list of (session_meta, results)."""
    runs = []
    for i in range(1, n_runs + 1):
        print(f"\n{'#'*72}")
        print(f"  RUN {i}/{n_runs}")
        print(f"{'#'*72}")
        csv_path = os.path.join(results_dir, f"test_results_run{i}.csv")
        meta, results = execute_campaign(
            sequence_name=sequence,
            unit_sn=f"{unit_sn}-R{i}",
            operator=operator,
            results_csv=csv_path,
            verbose=True,
        )
        runs.append((meta, results))
        if i < n_runs:
            print(f"\n[multi]  Pausing {inter_run_pause_s:.1f} s before run {i+1}…")
            time.sleep(inter_run_pause_s)
    return runs


# ── comparison workbook ───────────────────────────────────────────────────

def build_comparison_workbook(runs, out_path, sequence_name):
    n_runs = len(runs)
    test_keys = [r["test_key"] for r in runs[0][1]]

    wb = Workbook()

    # ──── Summary sheet ────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.cell(row=1, column=1,
            value=f"Multi-run HW PCBA Test Campaign — {n_runs} runs of "
                  f"sequence '{sequence_name}'").font = Font(
        name="Arial", size=14, bold=True, color=NAVY)
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=8)

    _hdr_row(ws, 3,
             ["Run", "Unit S/N", "Tests", "PASS", "FAIL",
              "TIMEOUT/ERROR", "Total Duration (s)", "Outcome"],
             widths=[8, 24, 8, 8, 8, 16, 18, 12])
    for i, (meta, results) in enumerate(runs, 1):
        row = 3 + i
        n_pass = sum(1 for r in results if r["result"] == "PASS")
        n_fail = sum(1 for r in results if r["result"] == "FAIL")
        n_oth  = len(results) - n_pass - n_fail
        total_s = sum(r["duration_s"] for r in results)
        outcome = "PASS" if (n_fail == 0 and n_oth == 0) else "FAIL"
        values = [i, meta["unit_sn"], len(results), n_pass, n_fail, n_oth,
                  round(total_s, 2), outcome]
        for ci, v in enumerate(values, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.alignment = CENTER
            c.border    = BORDER
            if ci == 8:
                c.fill = _result_fill(outcome)
                c.font = _result_font(outcome)
            elif ci == 4:
                c.fill = _fill(PASS_F)
            elif ci == 5 and n_fail > 0:
                c.fill = _fill(FAIL_F)
            elif ci == 6 and n_oth > 0:
                c.fill = _fill(WARN_F)
        ws.row_dimensions[row].height = 20

    # Aggregate row
    agg_row = 3 + n_runs + 2
    ws.cell(row=agg_row, column=1, value="Aggregate").font = Font(
        name="Arial", size=11, bold=True, color=NAVY)
    _hdr_row(ws, agg_row + 1,
             ["Metric", "Value"],
             widths=[36, 18])
    total_tests = n_runs * len(test_keys)
    total_pass  = sum(sum(1 for r in res if r["result"] == "PASS")
                     for _, res in runs)
    total_fail  = sum(sum(1 for r in res if r["result"] == "FAIL")
                     for _, res in runs)
    aggregate = [
        ("Total test executions",          total_tests),
        ("Total PASS",                     total_pass),
        ("Total FAIL",                     total_fail),
        ("Mean PASS rate (%)",             round(100 * total_pass / total_tests, 1)),
        ("Total wall-clock (s)",           round(sum(r["duration_s"]
                                                     for _, res in runs
                                                     for r in res), 2)),
    ]
    for i, (k, v) in enumerate(aggregate):
        r = agg_row + 2 + i
        ws.cell(row=r, column=1, value=k).border = BORDER
        ws.cell(row=r, column=1).alignment = LEFT
        ws.cell(row=r, column=1).font = Font(name="Arial", size=10, bold=True)
        ws.cell(row=r, column=2, value=v).border = BORDER
        ws.cell(row=r, column=2).alignment = CENTER

    # ──── Repeatability sheet ──────────────────────────────────────────
    ws = wb.create_sheet("Repeatability")
    ws.cell(row=1, column=1,
            value="Per-test repeatability across runs").font = Font(
        name="Arial", size=14, bold=True, color=NAVY)
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=n_runs + 5)

    headers = (["#", "Test Key", "Code"]
               + [f"Run {i}" for i in range(1, n_runs + 1)]
               + ["PASS rate", "Class.", "Notes"])
    widths = [5, 36, 7] + [10] * n_runs + [12, 14, 60]
    _hdr_row(ws, 3, headers, widths=widths)

    for ti, test_key in enumerate(test_keys, 1):
        row = 3 + ti
        results_for_test = [next(r for r in res if r["test_key"] == test_key)
                             for _, res in runs]
        pass_count = sum(1 for r in results_for_test if r["result"] == "PASS")
        pass_rate  = pass_count / n_runs

        if pass_count == n_runs:
            classification = "STABLE PASS"
            fill = _fill(PASS_F)
            font_cls = _result_font("PASS")
        elif pass_count == 0:
            classification = "STABLE FAIL"
            fill = _fill(FAIL_F)
            font_cls = _result_font("FAIL")
        else:
            classification = "FLAKY"
            fill = _fill(WARN_F)
            font_cls = _result_font("FLAKY")

        # Build per-row notes
        durs = [r["duration_s"] for r in results_for_test]
        notes_parts = [f"avg {sum(durs)/n_runs:5.2f}s, "
                        f"range {min(durs):5.2f}-{max(durs):5.2f}s"]
        flags = [r["flags_str"] for r in results_for_test]
        if len(set(flags)) > 1:
            notes_parts.append(f"flags differ: {flags}")
        # Truncate messages
        msgs = [r["message"][:60] for r in results_for_test
                if r["result"] != "PASS"]
        if msgs:
            unique_msgs = list(dict.fromkeys(msgs))[:2]
            notes_parts.append(" | ".join(unique_msgs))
        notes = "; ".join(notes_parts)

        # Values
        test_code_hex = f"0x{results_for_test[0]['test_code']:02X}"
        per_run_results = [r["result"] for r in results_for_test]
        values = ([ti, test_key.replace("HW_TEST_", ""), test_code_hex]
                  + per_run_results
                  + [f"{pass_rate*100:.0f}%", classification, notes])

        for ci, v in enumerate(values, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.border = BORDER
            c.alignment = LEFT if ci in (2, len(values)) else CENTER
            # Colour each Run column cell by its result
            if 4 <= ci <= 3 + n_runs:
                c.fill = _result_fill(v)
                c.font = _result_font(v)
            # Classification column
            elif ci == len(values) - 1:
                c.fill = fill
                c.font = font_cls
        ws.row_dimensions[row].height = 22

    ws.freeze_panes = "D4"

    # ──── Failure Diagnostics sheet ───────────────────────────────────
    ws = wb.create_sheet("Failure Diagnostics")
    ws.cell(row=1, column=1,
            value=("Analog readings on FAILED tests — one row per "
                   "(test, channel), values across runs side-by-side")
            ).font = Font(name="Arial", size=14, bold=True, color=NAVY)
    ws.merge_cells(start_row=1, end_row=1,
                   start_column=1, end_column=6 + n_runs)

    # Build the union of (test_key, ch_id) seen on any failing run, and
    # collect each run's value for that pair.
    failed_test_keys = set()
    for _meta, results in runs:
        for r in results:
            if r["result"] != "PASS":
                failed_test_keys.add(r["test_key"])

    headers = (["Test", "Code", "CH", "Name", "Unit", "Limit"]
               + [f"Run {i}" for i in range(1, n_runs + 1)])
    widths = [36, 7, 8, 26, 6, 22] + [12] * n_runs
    _hdr_row(ws, 3, headers, widths=widths)

    cur_row = 4
    if not failed_test_keys:
        ws.cell(row=cur_row, column=1, value="No failed tests across any run.")
        ws.merge_cells(start_row=cur_row, end_row=cur_row,
                       start_column=1, end_column=6 + n_runs)
    else:
        # Process failed tests in their sequence order
        ordered_keys = [r["test_key"] for r in runs[0][1]
                        if r["test_key"] in failed_test_keys]

        for test_key in ordered_keys:
            # Gather per-run results for this test
            per_run = [next(r for r in res if r["test_key"] == test_key)
                       for _, res in runs]
            code_hex = f"0x{per_run[0]['test_code']:02X}"

            # Header band per test (gentle navy)
            for ci in range(1, 6 + n_runs + 1):
                cc = ws.cell(row=cur_row, column=ci)
                cc.fill = _fill("D9E1F2")
                cc.border = BORDER
            ws.cell(row=cur_row, column=1,
                    value=f"{test_key.replace('HW_TEST_','')}  ({code_hex})"
                    ).font = Font(name="Arial", size=10, bold=True, color=NAVY)
            for i, r in enumerate(per_run, 7):
                vc = ws.cell(row=cur_row, column=i,
                             value=f"{r['result']}  ({r['flags_str']})")
                vc.alignment = CENTER
                vc.font = _result_font(r["result"])
            ws.row_dimensions[cur_row].height = 22
            cur_row += 1

            # Union of all analog channels reported by any run for this test
            all_chs = set()
            for r in per_run:
                all_chs.update(r["analog_values"].keys())

            if not all_chs:
                ws.cell(row=cur_row, column=1,
                        value="(no ANLG[] frames received — flag-only failure)"
                        ).font = Font(name="Arial", size=9, italic=True,
                                      color="808080")
                ws.merge_cells(start_row=cur_row, end_row=cur_row,
                               start_column=1, end_column=6 + n_runs)
                cur_row += 1
                continue

            for ch_id in sorted(all_chs):
                # Limits / unit from criteria
                unit, lim_str = "—", "no limit"
                if ch_id in ANLG_LIMITS:
                    lo, hi, u, _desc = ANLG_LIMITS[ch_id]
                    unit = u
                    lo_s = f"{lo}" if lo is not None else "−∞"
                    hi_s = f"{hi}" if hi is not None else "+∞"
                    lim_str = f"[{lo_s}, {hi_s}] {unit}"

                # Static columns
                cells = [
                    "", "", f"ANLG[{ch_id}]", channel_name(ch_id),
                    unit, lim_str,
                ]
                # Per-run values
                run_values = []
                for r in per_run:
                    if ch_id in r["analog_values"]:
                        v = r["analog_values"][ch_id]
                        ok, _ = r["analog_checks"].get(ch_id, (True, ""))
                        run_values.append((round(float(v), 4), ok))
                    else:
                        run_values.append((None, True))

                # Write static cols
                for ci, v in enumerate(cells, 1):
                    c = ws.cell(row=cur_row, column=ci, value=v)
                    c.border = BORDER
                    c.alignment = LEFT if ci == 4 else CENTER
                    c.font = Font(name="Arial", size=10)
                # Per-run cols (with PASS/FAIL fill on the cell)
                for i, (val, ok) in enumerate(run_values, 7):
                    c = ws.cell(row=cur_row, column=i,
                                value=("—" if val is None else val))
                    c.border = BORDER
                    c.alignment = CENTER
                    if val is None:
                        c.font = Font(name="Arial", size=10, color="999999")
                    elif ok:
                        c.font = Font(name="Arial", size=10)
                    else:
                        c.fill = _fill(FAIL_F)
                        c.font = Font(name="Arial", size=10, bold=True,
                                       color="9C0006")
                cur_row += 1

            cur_row += 1  # blank row between tests

    ws.freeze_panes = "G4"

    # ──── Per-run detail sheets ────────────────────────────────────────
    for i, (meta, results) in enumerate(runs, 1):
        ws = wb.create_sheet(f"Run {i}")
        ws.cell(row=1, column=1,
                value=f"Run {i} — {meta['unit_sn']} — "
                      f"{meta['test_date']}").font = Font(
            name="Arial", size=12, bold=True, color=NAVY)
        ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=8)

        _hdr_row(ws, 3,
                 ["#", "Test Key", "Code", "Result", "Flags",
                  "Dur (s)", "Analog", "Message"],
                 widths=[5, 38, 7, 10, 14, 9, 12, 60])
        for ti, r in enumerate(results, 1):
            row = 3 + ti
            values = [ti, r["test_key"].replace("HW_TEST_", ""),
                      f"0x{r['test_code']:02X}",
                      r["result"], r["flags_str"],
                      round(r["duration_s"], 2),
                      r["analog_result"], r["message"]]
            font_p = Font(name="Arial", size=10)
            font_b = _result_font(r["result"])
            fill = _result_fill(r["result"])
            for ci, v in enumerate(values, 1):
                c = ws.cell(row=row, column=ci, value=v)
                c.border = BORDER
                c.fill = fill
                c.font = font_b if ci == 4 else font_p
                c.alignment = LEFT if ci in (2, 8) else CENTER
        ws.freeze_panes = "A4"

    wb.save(out_path)
    return out_path


# ── CLI ───────────────────────────────────────────────────────────────────

def main():
    p = argparse.ArgumentParser(
        description="Run the HW PCBA test campaign N times and produce a "
                    "repeatability comparison report.")
    p.add_argument("--runs", "-n", type=int, default=3,
                   help="Number of campaign runs (default: 3)")
    p.add_argument("--sequence", "-s", default=TEST_SEQUENCE,
                   choices=["A0_A1", "A0", "A1", "B0", "B1", "B2"])
    p.add_argument("--unit-sn", "-u", default=UNIT_SN,
                   help="Base unit S/N (each run gets -R<n> suffix)")
    p.add_argument("--operator", "-o", default=OPERATOR)
    p.add_argument("--pause", type=float, default=2.0,
                   help="Pause between runs in seconds (default: 2.0)")
    args = p.parse_args()

    code_dir = os.path.dirname(os.path.abspath(__file__))
    results_dir = os.path.abspath(os.path.join(code_dir, "..", "results"))
    os.makedirs(results_dir, exist_ok=True)

    t_start = time.monotonic()
    runs = run_n_campaigns(args.runs, args.sequence, args.unit_sn,
                            args.operator, results_dir, args.pause)
    t_total = time.monotonic() - t_start

    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M")
    sn_safe = args.unit_sn.replace("/", "-").replace(" ", "_")
    out_xlsx = os.path.join(results_dir,
                             f"ValidationReport_{sn_safe}_{ts}_multi.xlsx")
    build_comparison_workbook(runs, out_xlsx, args.sequence)

    print(f"\n{'='*72}")
    print(f"  MULTI-RUN COMPLETE")
    print(f"  Total wall-clock : {t_total:.1f} s for {args.runs} runs")
    print(f"  Comparison xlsx  : {out_xlsx}")
    print(f"{'='*72}\n")
    return 0


if __name__ == "__main__":
    sys.exit(main())
