"""
generate_report.py  –  Stage 1 HW PCBA Validation Report generator (Excel).

Reads the results produced by hw_test_runner.execute_campaign() and writes a
styled multi-sheet Excel workbook (.xlsx).

Sheets
──────
  Summary         — cover info, executive PASS/FAIL counts, overall verdict
  Test Results    — one row per test with firmware flags + per-test result
  Analog Readings — every ANLG[] reading received, with limit check
  Conclusions     — overall verdict + engineer-notes area + sign-off block
"""

import os
import sys
import datetime
import statistics
from typing import List, Dict, Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment, Border, Side, Font, PatternFill,
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed.  Run:  pip install openpyxl")
    sys.exit(1)

from hw_test_criteria import ANLG_LIMITS, channel_name


# ── Colour palette ─────────────────────────────────────────────────────────
NAVY   = "1F3864"
WHITE  = "FFFFFF"
PASS_F = "E2EFDA"   # light green
FAIL_F = "FFE0DC"   # light red
WARN_F = "FFF2CC"   # light amber
HDR_F  = NAVY
SUB_F  = "D9E1F2"   # light blue
STRIPE = "F2F2F2"

PASS_TXT = "006100"
FAIL_TXT = "9C0006"
WARN_TXT = "9C6500"
NAVY_TXT = NAVY

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


# ── helpers ────────────────────────────────────────────────────────────────

def _fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")

def _result_fill(result_str: str):
    return _fill({"PASS": PASS_F,
                  "FAIL": FAIL_F,
                  "TIMEOUT": WARN_F,
                  "ERROR": WARN_F}.get(result_str, STRIPE))

def _result_font(result_str: str, *, bold=True):
    color = {"PASS": PASS_TXT, "FAIL": FAIL_TXT}.get(result_str, WARN_TXT)
    return Font(name="Arial", size=10, bold=bold, color=color)

def _write_row(ws, row_idx, values, *,
               fills=None, fonts=None, aligns=None, borders=True):
    for ci, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=ci, value=val)
        if fills and fills[ci - 1] is not None:
            cell.fill = fills[ci - 1]
        if fonts and fonts[ci - 1] is not None:
            cell.font = fonts[ci - 1]
        if aligns and aligns[ci - 1] is not None:
            cell.alignment = aligns[ci - 1]
        else:
            cell.alignment = LEFT
        if borders:
            cell.border = BORDER

def _header_row(ws, row_idx, labels, widths=None):
    hdr_font  = Font(name="Arial", size=10, bold=True, color=WHITE)
    hdr_fill  = _fill(HDR_F)
    for ci, lbl in enumerate(labels, 1):
        cell = ws.cell(row=row_idx, column=ci, value=lbl)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = CENTER
        cell.border    = BORDER
        if widths and ci - 1 < len(widths):
            ws.column_dimensions[get_column_letter(ci)].width = widths[ci - 1]
    ws.row_dimensions[row_idx].height = 22

def _section_title(ws, row_idx, text, span_cols):
    cell = ws.cell(row=row_idx, column=1, value=text)
    cell.font      = Font(name="Arial", size=14, bold=True, color=NAVY_TXT)
    cell.alignment = LEFT
    ws.merge_cells(start_row=row_idx, end_row=row_idx,
                   start_column=1, end_column=span_cols)
    ws.row_dimensions[row_idx].height = 24


# ── Sheets ─────────────────────────────────────────────────────────────────

def _build_summary_sheet(ws, results, session_meta, overall, counts):
    n_total, n_pass, n_fail, n_timeout, n_error = counts

    ws.title = "Summary"
    _section_title(ws, 1, "Inverter Gen3 — Stage 1 HW PCBA Validation Report", 6)

    info_rows = [
        ("Unit S/N",       session_meta.get("unit_sn",      "—")),
        ("Operator",       session_meta.get("operator",     "—")),
        ("Test sequence",  session_meta.get("sequence",     "—")),
        ("Test date",      session_meta.get("test_date",    "—")),
        ("CAN interface",  f"{session_meta.get('can_interface','—')}  "
                            f"{session_meta.get('can_channel','—')}  "
                            f"{session_meta.get('can_bitrate','—')} bps"),
        ("Overall result", overall),
    ]
    r = 3
    for label, value in info_rows:
        is_verdict = (label == "Overall result")
        ws.cell(row=r, column=1, value=label).font = Font(name="Arial", size=10, bold=True)
        ws.cell(row=r, column=1).fill = _fill(SUB_F)
        ws.cell(row=r, column=1).border = BORDER
        ws.cell(row=r, column=1).alignment = LEFT

        vc = ws.cell(row=r, column=2, value=value)
        vc.border = BORDER
        vc.alignment = LEFT
        if is_verdict:
            vc.fill = _result_fill(overall)
            vc.font = _result_font(overall, bold=True)
        else:
            vc.font = Font(name="Arial", size=10)
        ws.merge_cells(start_row=r, end_row=r, start_column=2, end_column=6)
        r += 1

    ws.column_dimensions["A"].width = 22
    for col in "BCDEF":
        ws.column_dimensions[col].width = 18

    # — Executive counts table
    r += 1
    _section_title(ws, r, "Executive Summary", 6); r += 1
    _header_row(ws, r, ["Tests Run", "PASS", "FAIL", "TIMEOUT", "ERROR", "Overall"])
    r += 1
    fills = [_fill(STRIPE),
             _fill(PASS_F),
             _fill(FAIL_F) if n_fail else _fill(STRIPE),
             _fill(WARN_F) if n_timeout else _fill(STRIPE),
             _fill(WARN_F) if n_error else _fill(STRIPE),
             _result_fill(overall)]
    fonts = [Font(name="Arial", size=10),
             Font(name="Arial", size=10, bold=True, color=PASS_TXT),
             Font(name="Arial", size=10, bold=True,
                  color=FAIL_TXT if n_fail else "000000"),
             Font(name="Arial", size=10, bold=True,
                  color=WARN_TXT if n_timeout else "000000"),
             Font(name="Arial", size=10, bold=True,
                  color=WARN_TXT if n_error else "000000"),
             _result_font(overall)]
    _write_row(ws, r,
               [n_total, n_pass, n_fail, n_timeout, n_error, overall],
               fills=fills,
               fonts=fonts,
               aligns=[CENTER] * 6)
    ws.row_dimensions[r].height = 22


def _build_test_results_sheet(wb, results):
    ws = wb.create_sheet("Test Results")
    _section_title(ws, 1, "A. Test Results (one row per test, execution order)", 9)

    cols = ["#", "Test Key", "Code", "Description",
            "Timeout (ms)", "Duration (s)", "Flags",
            "Analog Limit Check", "Result", "Message"]
    widths = [5, 36, 7, 50, 12, 12, 14, 18, 11, 60]
    _header_row(ws, 3, cols, widths)

    for idx, rec in enumerate(results, 1):
        row_idx = 3 + idx
        fill    = _result_fill(rec["result"])
        font    = Font(name="Arial", size=10)
        bold_f  = _result_font(rec["result"])

        values = [
            idx,
            rec["test_key"].replace("HW_TEST_", ""),
            f"0x{rec['test_code']:02X}",
            rec["description"],
            rec["timeout_ms"],
            round(rec["duration_s"], 3),
            rec["flags_str"],
            rec["analog_result"],
            rec["result"],
            rec["message"],
        ]
        fonts  = [font, font, font, font, font, font, font, font, bold_f, font]
        fills  = [fill] * len(values)
        aligns = [CENTER, LEFT, CENTER, LEFT, CENTER, CENTER, CENTER,
                  CENTER, CENTER, LEFT]
        _write_row(ws, row_idx, values, fills=fills, fonts=fonts, aligns=aligns)
        ws.row_dimensions[row_idx].height = 26

    ws.freeze_panes = "A4"


def _build_analog_sheet(wb, results):
    ws = wb.create_sheet("Analog Readings")
    _section_title(ws, 1, "B. Analog Channel Readings", 7)

    cols = ["Test", "CH ID", "Name", "Value", "Unit", "Limit", "Result"]
    widths = [32, 10, 28, 12, 8, 22, 11]
    _header_row(ws, 3, cols, widths)

    rows_data = []
    for rec in results:
        if not rec["analog_values"]:
            continue
        for ch_id, value in sorted(rec["analog_values"].items()):
            passed, msg_txt = rec["analog_checks"].get(
                ch_id, (True, "No limit defined"))
            rows_data.append((rec["test_key"].replace("HW_TEST_", ""),
                               ch_id, channel_name(ch_id), value, passed))

    if not rows_data:
        ws.cell(row=4, column=1, value="No analog readings collected.")
        ws.merge_cells(start_row=4, end_row=4, start_column=1, end_column=7)
        return

    for i, (test_key, ch_id, name, value, passed) in enumerate(rows_data, 1):
        row_idx = 3 + i
        result_str = "PASS" if passed else "FAIL"
        fill = _result_fill(result_str)
        font = Font(name="Arial", size=10)
        bold_f = _result_font(result_str)

        unit, lim = "—", "no limit"
        if ch_id in ANLG_LIMITS:
            lo, hi, unit_x, _desc = ANLG_LIMITS[ch_id]
            unit = unit_x
            lo_s = f"{lo}" if lo is not None else "−∞"
            hi_s = f"{hi}" if hi is not None else "+∞"
            lim  = f"[{lo_s}, {hi_s}] {unit}"

        values = [test_key, f"ANLG[{ch_id}]", name,
                  round(float(value), 4), unit, lim, result_str]
        fonts  = [font, font, font, font, font, font, bold_f]
        fills  = [fill] * 7
        aligns = [LEFT, CENTER, LEFT, CENTER, CENTER, CENTER, CENTER]
        _write_row(ws, row_idx, values, fills=fills, fonts=fonts, aligns=aligns)

    ws.freeze_panes = "A4"


def _build_conclusions_sheet(wb, results, session_meta, overall, counts):
    n_total, n_pass, n_fail, n_timeout, n_error = counts
    ws = wb.create_sheet("Conclusions")
    _section_title(ws, 1, "C. Conclusions", 3)

    r = 3
    ws.cell(row=r, column=1, value="Overall verdict:").font = Font(
        name="Arial", size=11, bold=True)
    ws.cell(row=r, column=1).alignment = LEFT
    vc = ws.cell(row=r, column=2, value=overall)
    vc.font = _result_font(overall, bold=True)
    vc.fill = _result_fill(overall)
    vc.alignment = CENTER
    vc.border = BORDER
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 80

    r += 2
    if overall == "PASS":
        body = (f"All {n_total} tests in sequence "
                f"'{session_meta.get('sequence','—')}' completed with PASS "
                f"status. No firmware verification failures (V−/S−) or analog "
                f"limit violations were detected. The board is considered "
                f"PCBA-assembly verified.")
    else:
        body = (f"The campaign completed with FAIL status. "
                f"{n_fail} test(s) failed, {n_timeout} timed out, "
                f"{n_error} error(s). Review the Test Results and Analog "
                f"Readings sheets for details. Rework or investigate the "
                f"failing functions before re-testing.")

    bc = ws.cell(row=r, column=1, value=body)
    bc.alignment = LEFT
    bc.font = Font(name="Arial", size=10)
    ws.merge_cells(start_row=r, end_row=r, start_column=1, end_column=3)
    ws.row_dimensions[r].height = 60

    # — Open notes / engineer comments
    r += 2
    ws.cell(row=r, column=1, value="Open notes / engineer comments:").font = (
        Font(name="Arial", size=10, bold=True))
    r += 1
    for _ in range(5):
        for ci in range(1, 4):
            ws.cell(row=r, column=ci, value="").border = BORDER
        ws.merge_cells(start_row=r, end_row=r, start_column=1, end_column=3)
        ws.row_dimensions[r].height = 24
        r += 1

    # — Sign-off block
    r += 2
    _section_title(ws, r, "Sign-Off", 3); r += 1
    _header_row(ws, r, ["Role", "Name", "Signature / Date"], widths=[24, 30, 36])
    r += 1
    roles = [
        ("Test Engineer", session_meta.get("operator", "—"), ""),
        ("HW Lead",        "",                                 ""),
        ("QA / Validation","",                                 ""),
    ]
    for role, name, sig in roles:
        vals = [role, name, sig]
        fills = [_fill(SUB_F), None, None]
        fonts = [Font(name="Arial", size=10, bold=True),
                 Font(name="Arial", size=10),
                 Font(name="Arial", size=10)]
        _write_row(ws, r, vals, fills=fills, fonts=fonts,
                   aligns=[LEFT, LEFT, LEFT])
        ws.row_dimensions[r].height = 24
        r += 1


# ── Gate-scope (manual scope-assisted) sheet ───────────────────────────────

def _fmt_scope(key, value):
    if value != value:           # NaN
        return "—"
    if key in ("rise_time", "fall_time"):
        return f"{value*1e9:.1f} ns"
    if key == "frequency":
        return f"{value/1000:.3f} kHz"
    if key == "duty":
        return f"{value:.2f} %"
    return f"{value:.3f} V"


def _build_gate_scope_sheet(wb, scope_records):
    """Add a 'Gate Scope Checks' sheet from a flat list of per-switch
    records, each: {switch, test, meas{}, checks{}, verdict}."""
    ws = wb.create_sheet("Gate Scope Checks")
    _section_title(ws, 1,
                   "D. Gate-signal scope verification (manual probe per switch)",
                   10)

    cols = ["Test", "Switch", "Setpoint", "Frequency", "Duty", "Rise", "Fall",
            "V-high", "V-low", "V-pp", "Verdict"]
    widths = [9, 8, 11, 13, 10, 10, 10, 10, 10, 10, 10]
    _header_row(ws, 3, cols, widths)

    keys = ["frequency", "duty", "rise_time", "fall_time",
            "v_high", "v_low", "v_pp"]
    # Measurement columns occupy spreadsheet columns 4..10 (after the 3
    # leading label columns Test/Switch/Setpoint); verdict is column 11.
    MEAS_FIRST_COL = 4
    VERDICT_COL    = 11
    row = 3
    for rec in scope_records:
        row += 1
        ok_all = rec.get("verdict", False)
        rfill = _fill(PASS_F if ok_all else FAIL_F)
        test_lbl = rec.get("test", "").upper()
        values = [test_lbl, rec.get("switch", ""), rec.get("setpoint", "—")] + \
                 [_fmt_scope(k, rec["meas"].get(k, float("nan"))) for k in keys] + \
                 ["PASS" if ok_all else "FAIL"]
        for ci, v in enumerate(values, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.border = BORDER
            c.alignment = CENTER if ci != 2 else LEFT
            if MEAS_FIRST_COL <= ci <= MEAS_FIRST_COL + 6:
                meas_key = keys[ci - MEAS_FIRST_COL]
                pok = rec.get("checks", {}).get(meas_key, (True, ""))[0]
                c.fill = _fill(PASS_F if pok else FAIL_F)
                if not pok:
                    c.font = Font(name="Arial", size=10, bold=True, color=FAIL_TXT)
            else:
                c.fill = rfill
            if ci == VERDICT_COL:
                c.font = Font(name="Arial", size=10, bold=True,
                              color=(PASS_TXT if ok_all else FAIL_TXT))
        ws.row_dimensions[row].height = 20
    ws.freeze_panes = "A4"


# ── Power-module sweep (manual 0x1B) sheet ─────────────────────────────────

# Standardised display: 2 decimals on every measured value (matches the
# standalone run_power_module_sweep.write_report).
_PM_DECIMALS = 2

def _fmt_pm(key, value):
    if value != value:                                   # NaN
        return "—"
    if key in ("rise_time", "fall_time"):
        return f"{value * 1e9:.{_PM_DECIMALS}f} ns"
    if key == "frequency":
        return f"{value / 1000.0:.{_PM_DECIMALS}f} kHz"
    if key == "duty":
        return f"{value:.{_PM_DECIMALS}f} %"
    return f"{value:.{_PM_DECIMALS}f} V"                 # voltages


def _build_power_module_sheet(wb, pm_records):
    """Add a 'Power Module Sweep' sheet from a flat list of per-setpoint
    records, each: {switch, freq_khz, duty_pct, deadband_ns, meas{}, checks{},
    verdict}. Produced by run_power_module_sweep.py (HW_TEST_ALL_POWER_MODULE
    0x1B), one row per (switch × frequency × duty) grid point.

    Polish matched to the standalone report: Arial on every cell, 2-decimal
    formatting, and three trailing summary rows (Mean / Std(abs) / Std(rel))
    computed per measured column. Frequency and Duty are swept setpoints so
    their aggregate stats are excluded (left blank)."""
    ws = wb.create_sheet("Power Module Sweep")
    _section_title(ws, 1,
                   "E. Power-module sweep verification "
                   "(HW_TEST_ALL_POWER_MODULE 0x1B, manual probe per switch)",
                   12)

    FONT = "Arial"
    F_DATA   = Font(name=FONT, size=10)
    F_SWITCH = Font(name=FONT, size=10, bold=True)
    F_FAIL   = Font(name=FONT, size=10, bold=True, color=FAIL_TXT)
    F_PASS_V = Font(name=FONT, size=10, bold=True, color=PASS_TXT)
    F_STAT_L = Font(name=FONT, size=10, bold=True)
    F_STAT_V = Font(name=FONT, size=10, bold=True, color=NAVY)

    cols = ["Switch", "Freq cmd", "Duty cmd", "DB (ns)", "Frequency", "Duty",
            "Rise", "Fall", "V-high", "V-low", "V-pp", "Verdict"]
    widths = [9, 10, 10, 8, 13, 10, 10, 10, 10, 10, 10, 10]
    _header_row(ws, 3, cols, widths)

    keys = ["frequency", "duty", "rise_time", "fall_time",
            "v_high", "v_low", "v_pp"]
    MEAS_FIRST_COL = 5
    VERDICT_COL    = 12
    row = 3
    for rec in pm_records:
        row += 1
        ok_all = rec.get("verdict", False)
        rfill = _fill(PASS_F if ok_all else FAIL_F)
        lead = [rec.get("switch", ""),
                f"{rec.get('freq_khz', float('nan')):.{_PM_DECIMALS}f} kHz",
                f"{rec.get('duty_pct', float('nan')):.{_PM_DECIMALS}f} %",
                f"{rec.get('deadband_ns', float('nan')):.0f}"]
        values = lead + \
                 [_fmt_pm(k, rec["meas"].get(k, float("nan"))) for k in keys] + \
                 ["PASS" if ok_all else "FAIL"]
        for ci, v in enumerate(values, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.border = BORDER
            c.alignment = CENTER if ci != 1 else LEFT
            if MEAS_FIRST_COL <= ci <= MEAS_FIRST_COL + 6:
                meas_key = keys[ci - MEAS_FIRST_COL]
                pok = rec.get("checks", {}).get(meas_key, (True, ""))[0]
                c.fill = _fill(PASS_F if pok else FAIL_F)
                c.font = F_DATA if pok else F_FAIL
            else:
                c.fill = rfill
                c.font = F_SWITCH if ci == 1 else F_DATA
            if ci == VERDICT_COL:
                c.font = F_PASS_V if ok_all else F_FAIL
        ws.row_dimensions[row].height = 20
    ws.freeze_panes = "A4"

    # ── summary rows (Mean / Std abs / Std rel) — freq & duty excluded ──────
    STATS_EXCLUDE = {"frequency", "duty"}
    SUM_FILL = _fill(SUB_F)
    series = {k: [rec["meas"].get(k, float("nan")) for rec in pm_records]
              for k in keys}
    series = {k: [v for v in vals if v == v] for k, vals in series.items()}

    def _stat(key, kind):
        vals = series.get(key, [])
        if not vals:
            return "—"
        mean = sum(vals) / len(vals)
        if kind == "mean":
            return _fmt_pm(key, mean)
        sd = statistics.pstdev(vals) if len(vals) > 1 else 0.0
        if kind == "abs":
            return _fmt_pm(key, sd)
        rel = (sd / abs(mean) * 100.0) if mean else float("nan")
        return "—" if rel != rel else f"{rel:.{_PM_DECIMALS}f} %"

    for label, kind in [("Mean", "mean"), ("Std (abs)", "abs"), ("Std (rel)", "rel")]:
        row += 1
        lc = ws.cell(row=row, column=1, value=label)
        lc.fill = SUM_FILL; lc.border = BORDER; lc.alignment = CENTER; lc.font = F_STAT_L
        for ci in (2, 3, 4, VERDICT_COL):
            bc = ws.cell(row=row, column=ci, value="")
            bc.fill = SUM_FILL; bc.border = BORDER
        for i, key in enumerate(keys):
            val = "" if key in STATS_EXCLUDE else _stat(key, kind)
            cell = ws.cell(row=row, column=MEAS_FIRST_COL + i, value=val)
            cell.fill = SUM_FILL; cell.border = BORDER
            cell.alignment = CENTER; cell.font = F_STAT_V


# ── Public API ─────────────────────────────────────────────────────────────

def generate_validation_report(results:      List[Dict[str, Any]],
                                session_meta: Dict[str, Any],
                                output_path:  str,
                                scope_records: List[Dict[str, Any]] = None,
                                pm_records:    List[Dict[str, Any]] = None) -> str:
    """
    Generate an Excel (.xlsx) Validation Report — one workbook for the whole
    PCBA verification.

    Parameters
    ----------
    results       : list of result dicts from hw_test_runner.run_sequence()
    session_meta  : dict produced by execute_campaign()
    output_path   : full path for the output .xlsx file
                    (if the path ends in .docx it is auto-rewritten to .xlsx)
    scope_records : optional list of per-switch gate-scope records (manual
                    scope-assisted checks). When provided, a "Gate Scope
                    Checks" sheet is added so the entire PCBA verification —
                    automated + manual — lives in a single file.
    pm_records    : optional list of per-setpoint power-module sweep records
                    (manual HW_TEST_ALL_POWER_MODULE 0x1B). When provided, a
                    "Power Module Sweep" sheet is added to the same workbook.
    """
    if output_path.lower().endswith(".docx"):
        output_path = output_path[:-5] + ".xlsx"

    n_total   = len(results)
    n_pass    = sum(1 for r in results if r["result"] == "PASS")
    n_fail    = sum(1 for r in results if r["result"] == "FAIL")
    n_timeout = sum(1 for r in results if r["result"] == "TIMEOUT")
    n_error   = sum(1 for r in results if r["result"] == "ERROR")
    overall   = "PASS" if (n_fail == 0 and n_timeout == 0 and n_error == 0) else "FAIL"
    counts    = (n_total, n_pass, n_fail, n_timeout, n_error)

    wb = Workbook()
    _build_summary_sheet(wb.active, results, session_meta, overall, counts)
    _build_test_results_sheet(wb, results)
    _build_analog_sheet(wb, results)
    if scope_records:
        _build_gate_scope_sheet(wb, scope_records)
    if pm_records:
        _build_power_module_sheet(wb, pm_records)
    _build_conclusions_sheet(wb, results, session_meta, overall, counts)

    # — Footer text on every sheet
    ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for ws in wb.worksheets:
        ws.oddFooter.center.text = (
            f"Generated {ts} — Inverter Gen3 Stage 1 HW PCBA Validation")
        ws.oddFooter.center.size = 8
        ws.oddFooter.center.color = "808080"

    out_dir = os.path.dirname(os.path.abspath(output_path))
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    wb.save(output_path)
    return os.path.abspath(output_path)
