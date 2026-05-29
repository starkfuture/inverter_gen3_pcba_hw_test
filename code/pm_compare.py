#!/usr/bin/env python3
"""
pm_compare.py  –  build an Excel that compares TWO power-module sweep sessions
for the same switch, point-by-point, to spot run-to-run inconsistencies
(e.g. a firmware variable that didn't update on one run).

Usage
─────
  python pm_compare.py --sessions B1_r1 B1_r2 --switch UBOT
  python pm_compare.py --sessions B1_r1 B1_r2 --switch UBOT --out my_compare.xlsx
"""
import sys
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

import os
import pickle
import argparse
import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import run_power_module_sweep as pm

_RESULTS_DIR = os.path.abspath(
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "results"))

NAVY, PASS_F, FAIL_F, DIFF_F, SUBL = "1F4E79", "E2EFDA", "FFE0DC", "FFF2CC", "D9E1F2"
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

# run-to-run "match" tolerances
TOL_F_FRAC = 0.02     # 2 % of commanded frequency
TOL_D_PP   = 1.5      # 1.5 percentage-points
TOL_V      = 0.20     # 0.20 V

FONT = "Arial"
F_TITLE = Font(name=FONT, size=14, bold=True, color=NAVY)
F_META_L = Font(name=FONT, size=10, bold=True)
F_META_V = Font(name=FONT, size=10)
F_HDR = Font(name=FONT, size=10, bold=True, color="FFFFFF")
F_DATA = Font(name=FONT, size=10)
F_DIFF = Font(name=FONT, size=10, bold=True, color="9C6500")
F_OK = Font(name=FONT, size=10, bold=True, color="006100")
F_BAD = Font(name=FONT, size=10, bold=True, color="9C0006")


def _fill(c):
    return PatternFill(start_color=c, end_color=c, fill_type="solid")


def _load(session, switch):
    p = os.path.join(_RESULTS_DIR, f".pm_session_{session}.pkl")
    if not os.path.exists(p):
        return [], {}
    data = pickle.load(open(p, "rb"))
    return data["switches"].get(switch, []), data.get("meta", {})


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--sessions", nargs=2, required=True, metavar=("A", "B"))
    ap.add_argument("--switch", default="UBOT")
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    r1, m1 = _load(args.sessions[0], args.switch)
    r2, m2 = _load(args.sessions[1], args.switch)
    if not r1 or not r2:
        print("Missing data — check session names / switch.")
        return 1
    n = min(len(r1), len(r2))

    wb = Workbook(); ws = wb.active; ws.title = f"{args.switch} run compare"
    NCOL = 17
    tc = ws.cell(row=1, column=1,
                 value=f"Inverter Gen3 — Power-module sweep run comparison "
                       f"({args.switch}, HW_TEST_ALL_POWER_MODULE 0x1B)")
    tc.font = F_TITLE
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=NCOL)

    r = 3
    for k, v in [("Switch", args.switch),
                 ("Run 1 (session)", args.sessions[0]),
                 ("Run 2 (session)", args.sessions[1]),
                 ("Scope", m1.get("scope_idn", "—")),
                 ("Dead-band", m1.get("deadband", "—")),
                 ("Match tolerances", f"freq ±{TOL_F_FRAC*100:g}% · duty ±{TOL_D_PP:g}pp · V ±{TOL_V:g}V"),
                 ("Generated", datetime.datetime.now().isoformat(timespec="seconds"))]:
        lc = ws.cell(row=r, column=1, value=k); lc.font = F_META_L
        lc.fill = _fill(SUBL); lc.border = BORDER
        vc = ws.cell(row=r, column=2, value=str(v)); vc.font = F_META_V; vc.border = BORDER
        ws.merge_cells(start_row=r, end_row=r, start_column=2, end_column=NCOL)
        r += 1

    # grouped header
    r += 1
    cols = ["Freq cmd", "Duty cmd",
            "F r1 (kHz)", "F r2 (kHz)", "ΔF (kHz)",
            "Duty r1 (%)", "Duty r2 (%)", "ΔDuty (pp)",
            "Vhi r1 (V)", "Vhi r2 (V)",
            "Vlo r1 (V)", "Vlo r2 (V)",
            "Rise r1 (ns)", "Rise r2 (ns)",
            "Fall r1 (ns)", "Fall r2 (ns)",
            "Match"]
    widths = [9, 9, 10, 10, 9, 10, 10, 10, 9, 9, 9, 9, 11, 11, 11, 11, 8]
    for ci, (lbl, w) in enumerate(zip(cols, widths), 1):
        c = ws.cell(row=r, column=ci, value=lbl)
        c.font = F_HDR; c.fill = _fill(NAVY); c.alignment = CENTER; c.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    hdr = r

    def gv(rec, key):
        v = rec["meas"].get(key, float("nan"))
        return float("nan") if v != v else v

    n_mismatch = 0
    for i in range(n):
        a, b = r1[i], r2[i]
        r += 1
        fc = a["freq_khz"]; dc = a["duty_pct"]
        fa, fb = gv(a, "frequency") / 1000.0, gv(b, "frequency") / 1000.0
        da, db = gv(a, "duty"), gv(b, "duty")
        vha, vhb = gv(a, "v_high"), gv(b, "v_high")
        vla, vlb = gv(a, "v_low"), gv(b, "v_low")
        ra, rb = gv(a, "rise_time") * 1e9, gv(b, "rise_time") * 1e9
        flla, fllb = gv(a, "fall_time") * 1e9, gv(b, "fall_time") * 1e9
        dF = fb - fa; dD = db - da
        # match if every run-to-run delta is within tolerance (and not NaN)
        vals = [fa, fb, da, db, vha, vhb, vla, vlb]
        nan_any = any(v != v for v in vals)
        match = (not nan_any
                 and abs(dF) <= TOL_F_FRAC * max(fc, 1)
                 and abs(dD) <= TOL_D_PP
                 and abs(vhb - vha) <= TOL_V
                 and abs(vlb - vla) <= TOL_V)
        if not match:
            n_mismatch += 1
        rowfill = _fill(PASS_F if match else DIFF_F)

        def put(ci, text, *, diff=False, font=None):
            c = ws.cell(row=r, column=ci, value=text)
            c.fill = _fill(DIFF_F) if diff else rowfill
            c.border = BORDER; c.alignment = CENTER
            c.font = font if font else (F_DIFF if diff else F_DATA)

        put(1, f"{fc:.2f} kHz"); put(2, f"{dc:.2f} %")
        put(3, f"{fa:.2f}"); put(4, f"{fb:.2f}")
        put(5, f"{dF:+.2f}", diff=abs(dF) > TOL_F_FRAC * max(fc, 1))
        put(6, f"{da:.2f}"); put(7, f"{db:.2f}")
        put(8, f"{dD:+.2f}", diff=abs(dD) > TOL_D_PP)
        put(9, f"{vha:.2f}"); put(10, f"{vhb:.2f}")
        put(11, f"{vla:.2f}"); put(12, f"{vlb:.2f}")
        put(13, f"{ra:.2f}"); put(14, f"{rb:.2f}")
        put(15, f"{flla:.2f}"); put(16, f"{fllb:.2f}")
        mc = ws.cell(row=r, column=17, value=("MATCH" if match else "DIFF"))
        mc.fill = rowfill; mc.border = BORDER; mc.alignment = CENTER
        mc.font = F_OK if match else F_BAD
    ws.freeze_panes = ws.cell(row=hdr + 1, column=3)

    out = args.out or os.path.join(
        _RESULTS_DIR,
        f"PowerModuleSweep_Compare_{args.switch}_"
        f"{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
    os.makedirs(os.path.dirname(os.path.abspath(out)), exist_ok=True)
    wb.save(out)
    print(f"Compared {n} points — {n - n_mismatch} MATCH, {n_mismatch} DIFF")
    print(f"Report: {os.path.abspath(out)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
