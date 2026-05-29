#!/usr/bin/env python3
"""
run_gate_scope_check.py  –  scope-assisted gate-signal verification for the
ALL_DRIVERS_FREQUENCY (0x18) and ALL_DRIVERS_DUTY (0x19) tests.

The firmware drives the 6 isolated gate PWM signals and reports only that no
fault is present and all drivers are ready. This tool adds the missing
electrical verification: with a voltage probe on one gate at a time, it uses
a Rohde & Schwarz touch scope (over USB) to measure, switch by switch:

    frequency, duty cycle, rise time, fall time, V-high, V-low, V-pp

Workflow (interactive, one switch at a time)
────────────────────────────────────────────
  For each of UTOP, UBOT, VTOP, VBOT, WTOP, WBOT:
    1. The tool asks you to move the probe to that gate and press ENTER.
    2. It commands the firmware to drive the PWM (optionally at a fixed
       frequency / duty setpoint) so the signal is stable.
    3. It configures the scope and captures the 7 parameters.
    4. It records them (and optionally a screenshot).
  Finally it writes results/GateScopeReport_<SN>_<ts>.xlsx.

Usage
─────
  # Frequency test, scope on CH1, fixed 16 kHz, +18/-3 gate, 10:1 probe:
  python run_gate_scope_check.py --test freq --scope-ch 1 --freq-khz 16

  # Duty test at 50 %:
  python run_gate_scope_check.py --test duty --duty 50

  # Dry-run with no hardware (canned scope + CAN skipped):
  python run_gate_scope_check.py --test freq --simulate

  # Explicit VISA resource if auto-discovery picks the wrong device:
  python run_gate_scope_check.py --resource "USB0::0x0AAD::0x01D6::123456::INSTR"

Prerequisites (at the bench)
────────────────────────────
  pip install pyvisa pyvisa-py pyusb
  PEAK PCAN-USB connected; board flashed with HW-test firmware.
"""
import sys
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

import os
import time
import argparse
import datetime
from typing import Dict, Any, List

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from hw_protocol import HwTestProtocol
from scope_rs import RohdeScope

# ── The 6 power switches, in inverter half-bridge order ───────────────────
SWITCHES = ["UTOP", "UBOT", "VTOP", "VBOT", "WTOP", "WBOT"]

# ── Default scope setup for +18 V / -3 V gate drive (10:1 probe) ──────────
GATE_V_HIGH_NOM   = 18.0
GATE_V_LOW_NOM    = -3.0
DEFAULT_PROBE_ATT = 10.0
DEFAULT_V_DIV     = 4.0     # 8 div × 4 V = 32 V window, centred on +7.5 V
DEFAULT_OFFSET    = 7.5     # midpoint of +18 / -3
DEFAULT_TRIG_LVL  = 5.0

# ── Soft pass/fail limits (report-only if a value is None) ────────────────
# (min, max, unit). Tune to your gate-driver spec.
PARAM_LIMITS = {
    "frequency": (None, None, "Hz"),   # checked against the commanded setpoint
    "duty":      (None, None, "%"),    # checked against the commanded setpoint
    "rise_time": (None, 500e-9, "s"),  # < 500 ns (loose default)
    "fall_time": (None, 500e-9, "s"),
    "v_high":    (16.0, 20.0, "V"),    # +18 V ± 2 V
    "v_low":     (-5.0, -1.0, "V"),    # -3 V ± 2 V
    "v_pp":      (18.0, 25.0, "V"),
}

# ── styles ────────────────────────────────────────────────────────────────
NAVY, PASS_F, FAIL_F, STRIPE = "1F4E79", "E2EFDA", "FFE0DC", "F2F2F2"
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def _fill(c): return PatternFill(start_color=c, end_color=c, fill_type="solid")


# ── CAN helpers (skipped in --simulate) ───────────────────────────────────

def _open_can():
    from hw_can_utils import HwTestCANBus
    bus = HwTestCANBus().open()
    return bus

def _send(bus, protocol, rqst, instance, value):
    from hw_can_utils import send_frame
    res = protocol.sfHwTestProtocolProcessTxMessage(rqst, instance, value)
    if res and res[0] > 0:
        send_frame(bus, res[1])


def _start_pwm_test(bus, protocol, which_test: str,
                    freq_khz: float = None, duty_pct: float = None):
    """Put the firmware into ALL_DRIVERS_FREQUENCY or ALL_DRIVERS_DUTY and,
    if a setpoint is given, force a fixed frequency / duty so the scope sees
    a stable waveform."""
    test_key = ("HW_TEST_ALL_DRIVERS_FREQUENCY" if which_test == "freq"
                else "HW_TEST_ALL_DRIVERS_DUTY")
    # SET_TEST_ENV (report enable + clear reboot)
    flags = (protocol.SF_HW_TEST_PROTOCOL_TEST_SET_ENV_FLAGS_REPORT_ENABLE
             | protocol.SF_HW_TEST_PROTOCOL_TEST_SET_ENV_FLAGS_CLEAR_REBOOT)
    _send(bus, protocol, "SF_HW_TEST_PROTOCOL_RQST_SET_TEST_ENV", 0, flags)
    time.sleep(0.3)
    _send(bus, protocol, "SF_HW_TEST_PROTOCOL_RQST_SET_TEST", 0, test_key)
    time.sleep(0.3)
    if freq_khz is not None:
        _send(bus, protocol, "SF_HW_TEST_PROTOCOL_RQST_SET_PWM_FREQ", 0, freq_khz)
        time.sleep(0.2)
    if duty_pct is not None:
        _send(bus, protocol, "SF_HW_TEST_PROTOCOL_RQST_SET_PWM_DUTY", 0, duty_pct)
        time.sleep(0.2)


def _stop_pwm_test(bus, protocol):
    _send(bus, protocol, "SF_HW_TEST_PROTOCOL_RQST_SET_TEST", 0, "HW_TEST_NO_TEST")
    time.sleep(0.1)


# ── parameter check ────────────────────────────────────────────────────────

def _check_param(key, value, setpoints=None):
    """Return (passed, note).

    *setpoints* is a dict like {"frequency": 16000.0} or {"duty": 50.0}
    holding ONLY the parameter that the firmware was commanded to (the
    swept quantity). That parameter is checked against its setpoint ±5 %;
    every other parameter is checked against PARAM_LIMITS.
    """
    if value != value:  # NaN
        return False, "no reading"
    setpoints = setpoints or {}
    if key in setpoints and setpoints[key]:
        sp = setpoints[key]
        tol = 0.05 * sp
        ok = abs(value - sp) <= tol
        return ok, f"setpoint {sp:g} ±5%"
    lo, hi, _unit = PARAM_LIMITS.get(key, (None, None, ""))
    if lo is not None and value < lo:
        return False, f"< {lo}"
    if hi is not None and value > hi:
        return False, f"> {hi}"
    return True, "ok"


def _fmt(key, value):
    if value != value:
        return "—"
    if key in ("rise_time", "fall_time"):
        return f"{value*1e9:.1f} ns"
    if key == "frequency":
        return f"{value/1000:.3f} kHz"
    if key == "duty":
        return f"{value:.2f} %"
    return f"{value:.3f} V"


# ── Sweep-plan + single-setpoint measurement (shared by this tool and
#    verify_pcba.py) ───────────────────────────────────────────────────────

# Defaults for range-verify. Duty endpoints avoid the degenerate 0 %/100 %
# (constant DC, no switching edges to measure); 0/100 can be added but only
# V-high/V-low are meaningful there.
DEFAULT_FREQ_POINTS = [10.0, 20.0, 30.0]   # kHz — matches FW sweep 10..30
DEFAULT_DUTY_POINTS = [10.0, 50.0, 90.0]   # %  — within FW sweep 0..100
DUTY_VERIFY_CARRIER_KHZ = 16.0             # fixed carrier while sweeping duty


def build_plan(test: str, range_verify: bool,
               freq_khz: float, duty: float,
               freq_points=None, duty_points=None):
    """Return an ordered list of setpoint dicts to measure per switch.

    Each entry: {test_type, freq_khz, duty, label, setpoints}
      test_type : 'freq' | 'duty'   (which firmware test drives the PWM)
      freq_khz  : frequency setpoint to command (or None = firmware sweep)
      duty      : duty setpoint to command (or None = firmware sweep)
      label     : human string for the report ("20 kHz", "50 %")
      setpoints : dict for the ±5 % check, e.g. {"frequency": 20000.0}
    """
    freq_points = freq_points or DEFAULT_FREQ_POINTS
    duty_points = duty_points or DEFAULT_DUTY_POINTS
    do_freq = test in ("freq", "both")
    do_duty = test in ("duty", "both")
    plan = []

    if range_verify:
        if do_freq:
            for f in freq_points:
                plan.append(dict(test_type="freq", freq_khz=f, duty=50.0,
                                 label=f"{f:g} kHz",
                                 setpoints={"frequency": f * 1000.0}))
        if do_duty:
            for d in duty_points:
                plan.append(dict(test_type="duty",
                                 freq_khz=DUTY_VERIFY_CARRIER_KHZ, duty=d,
                                 label=f"{d:g} %",
                                 setpoints={"duty": d}))
    else:
        if do_freq:
            plan.append(dict(test_type="freq", freq_khz=freq_khz, duty=50.0,
                             label=(f"{freq_khz:g} kHz" if freq_khz else "FW sweep"),
                             setpoints=({"frequency": freq_khz * 1000.0}
                                        if freq_khz else {})))
        if do_duty:
            plan.append(dict(test_type="duty",
                             freq_khz=DUTY_VERIFY_CARRIER_KHZ, duty=duty,
                             label=(f"{duty:g} %" if duty else "FW sweep"),
                             setpoints=({"duty": duty} if duty else {})))
    return plan


def gen_points(mn: float, mx: float, step: float):
    """Inclusive list of points from mn to mx in `step` increments."""
    if step <= 0 or mx < mn:
        return [mn]
    pts, v = [], mn
    while v <= mx + 1e-9:
        pts.append(round(v, 6))
        v += step
    if pts[-1] < mx - 1e-9:
        pts.append(mx)
    return pts


def measure_one_setpoint(scope, bus, protocol, scope_ch, entry, simulate,
                         edge_t_div: float = None):
    """Command one setpoint, measure the scope, return a result record.

    edge_t_div : fast horizontal scale for the accurate rise/fall second pass
                 (see RohdeScope.measure). None = single coarse pass only.
    """
    if not simulate:
        _start_pwm_test(bus, protocol, entry["test_type"],
                        freq_khz=entry["freq_khz"] if entry["test_type"] == "freq" else None,
                        duty_pct=entry["duty"] if entry["test_type"] == "duty" else None)
        # For freq sweeps, also pin the duty to 50 % so the waveform is
        # measurable; for duty sweeps, pin the carrier frequency.
        if entry["test_type"] == "freq" and entry.get("duty"):
            _send(bus, protocol, "SF_HW_TEST_PROTOCOL_RQST_SET_PWM_DUTY", 0, entry["duty"])
        if entry["test_type"] == "duty" and entry.get("freq_khz"):
            _send(bus, protocol, "SF_HW_TEST_PROTOCOL_RQST_SET_PWM_FREQ", 0, entry["freq_khz"])
    time.sleep(0.4)
    meas = scope.measure(scope_ch, edge_t_div=edge_t_div)
    checks = {k: _check_param(k, v, entry["setpoints"]) for k, v in meas.items()}
    verdict = all(ok for ok, _ in checks.values())
    if not simulate:
        _stop_pwm_test(bus, protocol)
    return dict(test=entry["test_type"], setpoint=entry["label"],
                meas=meas, checks=checks, verdict=verdict)


# ── report ─────────────────────────────────────────────────────────────────

def write_report(results: List[Dict[str, Any]], meta: Dict[str, Any],
                 out_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Gate Scope Check"

    ws.cell(row=1, column=1,
            value=f"Inverter Gen3 — Gate-signal scope check "
                  f"({meta['test'].upper()} sweep)").font = Font(
        name="Arial", size=14, bold=True, color=NAVY)
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=9)

    # meta block
    r = 3
    for k, v in [("Unit S/N", meta["unit_sn"]), ("Operator", meta["operator"]),
                 ("Scope", meta["scope_idn"]), ("Probe atten", meta["probe_atten"]),
                 ("Setpoint", meta["setpoint_str"]), ("Date", meta["date"])]:
        ws.cell(row=r, column=1, value=k).font = Font(name="Arial", bold=True, size=10)
        ws.cell(row=r, column=1).fill = _fill("D9E1F2")
        ws.cell(row=r, column=1).border = BORDER
        c = ws.cell(row=r, column=2, value=str(v)); c.border = BORDER
        ws.merge_cells(start_row=r, end_row=r, start_column=2, end_column=9)
        r += 1

    # header
    r += 1
    cols = ["Switch", "Test", "Setpoint", "Frequency", "Duty", "Rise", "Fall",
            "V-high", "V-low", "V-pp", "Verdict"]
    widths = [9, 8, 11, 13, 10, 10, 10, 10, 10, 10, 10]
    for ci, (lbl, w) in enumerate(zip(cols, widths), 1):
        c = ws.cell(row=r, column=ci, value=lbl)
        c.font = Font(name="Arial", bold=True, color="FFFFFF")
        c.fill = _fill(NAVY); c.alignment = CENTER; c.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    hdr_row = r

    keys = ["frequency", "duty", "rise_time", "fall_time", "v_high", "v_low", "v_pp"]
    MEAS_FIRST = 4   # measurement cols start at spreadsheet column 4
    for rec in results:
        r += 1
        verdict_ok = rec["verdict"]
        rowfill = _fill(PASS_F if verdict_ok else FAIL_F)
        # leading label cols
        for ci, val in enumerate([rec["switch"], rec.get("test", "").upper(),
                                   rec.get("setpoint", "—")], 1):
            cc = ws.cell(row=r, column=ci, value=val)
            cc.fill = rowfill; cc.border = BORDER
            cc.alignment = CENTER if ci != 1 else CENTER
            if ci == 1:
                cc.font = Font(name="Arial", bold=True)
        for i, key in enumerate(keys):
            ci = MEAS_FIRST + i
            val = rec["meas"].get(key, float("nan"))
            cell = ws.cell(row=r, column=ci, value=_fmt(key, val))
            ok = rec["checks"].get(key, (True, ""))[0]
            cell.fill = _fill(PASS_F if ok else FAIL_F)
            cell.border = BORDER; cell.alignment = CENTER
            if not ok:
                cell.font = Font(name="Arial", bold=True, color="9C0006")
        vc = ws.cell(row=r, column=11, value=("PASS" if verdict_ok else "FAIL"))
        vc.fill = rowfill; vc.border = BORDER; vc.alignment = CENTER
        vc.font = Font(name="Arial", bold=True,
                       color=("006100" if verdict_ok else "9C0006"))
    ws.freeze_panes = ws.cell(row=hdr_row + 1, column=1)

    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
    wb.save(out_path)
    return os.path.abspath(out_path)


# ── main ─────────────────────────────────────────────────────────────────

def main():
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--test", choices=["freq", "duty", "both"], required=True,
                   help="Which firmware test drives the PWM: freq=0x18, "
                        "duty=0x19, both=run both")
    p.add_argument("--range-verify", action="store_true",
                   help="Measure several defined points across the sweep range "
                        "instead of a single setpoint, so you confirm the whole "
                        "range. Frequency points come from --freq-points or the "
                        "--freq-min/max/step sweep; duty from --duty-points.")
    p.add_argument("--freq-points", nargs="+", type=float, default=None,
                   help="Explicit frequency points (kHz) for --range-verify")
    p.add_argument("--freq-min", type=float, default=10.0,
                   help="Frequency sweep min (kHz) when --freq-points not given (FW range 10)")
    p.add_argument("--freq-max", type=float, default=30.0,
                   help="Frequency sweep max (kHz) (FW range 30)")
    p.add_argument("--freq-step", type=float, default=5.0,
                   help="Frequency sweep step (kHz)")
    p.add_argument("--duty-points", nargs="+", type=float, default=None,
                   help="Duty points (%%) for --range-verify (default 10 50 90)")
    p.add_argument("--edge-tdiv", type=float, default=100e-9,
                   help="Fast timebase (s/div) for the accurate rise/fall "
                        "second pass (default 100e-9 = 100 ns/div). Set 0 to "
                        "disable edge-zoom.")
    p.add_argument("--scope-ch", type=int, default=1, help="Scope channel (default 1)")
    p.add_argument("--resource", default=None, help="Explicit VISA resource string")
    p.add_argument("--backend", default="@ivi",
                   help="VISA backend: '@ivi' (vendor VISA — required for USB "
                        "on this bench, default) or '@py' (pyvisa-py)")
    p.add_argument("--freq-khz", type=float, default=None,
                   help="Force PWM frequency setpoint (kHz)")
    p.add_argument("--duty", type=float, default=None,
                   help="Force PWM duty setpoint (%%)")
    p.add_argument("--probe-atten", type=float, default=DEFAULT_PROBE_ATT)
    p.add_argument("--v-div", type=float, default=DEFAULT_V_DIV)
    p.add_argument("--offset", type=float, default=DEFAULT_OFFSET)
    p.add_argument("--trig-level", type=float, default=DEFAULT_TRIG_LVL)
    p.add_argument("--screenshot", action="store_true",
                   help="Save a PNG screenshot per switch")
    p.add_argument("--unit-sn", default="PCB-B1")
    p.add_argument("--operator", default="Carlos Miguel Espinar")
    p.add_argument("--simulate", action="store_true",
                   help="No hardware: canned scope data, CAN skipped")
    p.add_argument("--switches", nargs="+", default=SWITCHES,
                   help="Subset/order of switches to check")
    args = p.parse_args()

    # Frequency points: explicit --freq-points, else generated from the
    # parametric --freq-min/max/step sweep (defaults to FW range 10..30 kHz).
    freq_points = args.freq_points
    if freq_points is None:
        freq_points = gen_points(args.freq_min, args.freq_max, args.freq_step)
    edge_t_div = args.edge_tdiv if args.edge_tdiv and args.edge_tdiv > 0 else None

    # Build the per-switch measurement plan (single setpoint or range-verify)
    plan = build_plan(args.test, args.range_verify,
                      args.freq_khz, args.duty,
                      freq_points, args.duty_points)
    if args.range_verify:
        setpoint_str = (f"range-verify  freq={args.freq_points or DEFAULT_FREQ_POINTS} kHz, "
                        f"duty={args.duty_points or DEFAULT_DUTY_POINTS} %")
    else:
        parts = [e["label"] for e in plan]
        setpoint_str = " + ".join(parts)

    def _tdiv_for(entry):
        f_hz = (entry["freq_khz"] * 1000.0) if entry["freq_khz"] else None
        if f_hz:
            return max((1.0 / f_hz) * 3 / 10.0, 1e-7)
        return 20e-6

    protocol = HwTestProtocol()
    bus = None if args.simulate else _open_can()
    scope = RohdeScope(resource=args.resource, backend=args.backend,
                       simulate=args.simulate).open()

    print("=" * 72)
    print(f"  Gate-signal scope check — {args.test.upper()}"
          f"{'  (range-verify)' if args.range_verify else ''}")
    print(f"  Scope     : {scope.idn()}")
    print(f"  Scope ch  : CH{args.scope_ch}   probe {args.probe_atten:g}:1   "
          f"{args.v_div} V/div, offset {args.offset} V")
    print(f"  Setpoints : {setpoint_str}")
    print(f"  Switches  : {', '.join(args.switches)}   "
          f"({len(plan)} setpoint(s) per switch)")
    print("=" * 72)

    # Full pre-configuration (reset + vertical + trigger + 7 on-screen
    # measurement slots). Timebase is then adjusted per setpoint below.
    scope.full_setup(args.scope_ch, v_div=args.v_div, offset=args.offset,
                     probe_atten=args.probe_atten, t_div=_tdiv_for(plan[0]),
                     trig_level=args.trig_level)

    results = []
    try:
        for sw in args.switches:
            print()
            print("  " + "═" * 68)
            print(f"  ⚠  Move the voltage probe to the  {sw}  gate (isolated PWM).")
            print("  " + "═" * 68)
            if not args.simulate:
                try:
                    input(f"  Press ENTER when the probe is on {sw} … ")
                except (EOFError, KeyboardInterrupt):
                    print("\n  Aborted by operator."); break

            for entry in plan:
                scope.configure_timebase(_tdiv_for(entry))
                rec = measure_one_setpoint(scope, bus, protocol,
                                           args.scope_ch, entry, args.simulate,
                                           edge_t_div=edge_t_div)
                rec["switch"] = sw
                rec["screenshot"] = ""
                if args.screenshot and not args.simulate:
                    shot_path = os.path.join(
                        os.path.dirname(os.path.abspath(__file__)), "..",
                        "results",
                        f"gate_{sw}_{entry['test_type']}_{entry['label'].replace(' ','')}.png")
                    if scope.screenshot(shot_path):
                        rec["screenshot"] = shot_path
                results.append(rec)
                m = rec["meas"]
                print(f"    [{entry['test_type']} @ {entry['label']:>7}]  "
                      f"f={_fmt('frequency', m['frequency'])}  "
                      f"duty={_fmt('duty', m['duty'])}  "
                      f"Vhi={_fmt('v_high', m['v_high'])}  "
                      f"Vlo={_fmt('v_low', m['v_low'])}  "
                      f"→ {'PASS' if rec['verdict'] else 'FAIL'}")
    finally:
        if not args.simulate:
            bus.close()
        scope.close()

    # Report (standalone gate-scope report; for the unified PCBA report use
    # verify_pcba.py --scope which folds these records into one workbook).
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M")
    sn_safe = args.unit_sn.replace("/", "-").replace(" ", "_")
    out = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                       "..", "results",
                       f"GateScopeReport_{sn_safe}_{args.test}_{ts}.xlsx")
    meta = dict(test=args.test, unit_sn=args.unit_sn, operator=args.operator,
                scope_idn=scope.idn(), probe_atten=f"{args.probe_atten:g}:1",
                setpoint_str=setpoint_str,
                date=datetime.datetime.now().isoformat(timespec="seconds"))
    out = write_report(results, meta, out)

    n_pass = sum(1 for r in results if r["verdict"])
    print()
    print("=" * 72)
    print(f"  GATE SCOPE CHECK COMPLETE — {n_pass}/{len(results)} switches PASS")
    print(f"  Report: {out}")
    print("=" * 72)
    return 0 if n_pass == len(results) else 1


if __name__ == "__main__":
    sys.exit(main())
