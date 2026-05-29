#!/usr/bin/env python3
"""
run_power_module_sweep.py  –  manual scope-assisted sweep of the power stage
using HW_TEST_ALL_POWER_MODULE (0x1B).

Unlike the gate-only checks (run_gate_scope_check.py, tests 0x18/0x19) this
drives ALL power branches simultaneously (the firmware HW_TEST_ALL_POWER_MODULE
test) and lets you sweep THREE quantities manually while a Rohde & Schwarz
scope measures the gate of one switch at a time:

    • switching frequency   (SET_PWM_FREQ 0x04, value in kHz)
    • duty cycle            (SET_PWM_DUTY 0x05, value PER-UNIT 0.0–1.0)
    • dead-band / deadtime  (SET_PWM_DT  0x07, value in ns)

For each of the 6 switches (UTOP, UBOT, VTOP, VBOT, WTOP, WBOT) the operator
moves the probe ONCE; the tool then walks the full frequency × duty grid at the
fixed dead-band and captures, per setpoint:

    frequency, duty, rise time, fall time, V-high, V-low, V-pp

────────────────────────────────────────────────────────────────────────────
Sweep notation
────────────────────────────────────────────────────────────────────────────
The default sweep is "10-10-30 kHz, 25-25-75 %, dead-band 1 µs":

    --freq-min 10  --freq-step 10 --freq-max 30     →  10,20,30 kHz
    --duty-min 25  --duty-step 25 --duty-max 75     →  25,50,75 %  (UI percent)
    --deadband-ns 1000                              →  1 µs

Duty is entered as PERCENT on the command line but sent to the firmware
PER-UNIT (25 %→0.25, 50 %→0.5, 75 %→0.75). Dead-band is entered/sent in ns.

⚠  This test energises the real power branches. Make sure the DC-link / power
   stage is in the intended (safe, low-voltage or de-energised gate-drive) state
   for your bench before running. Use --simulate to dry-run with canned data.

────────────────────────────────────────────────────────────────────────────
Usage
────────────────────────────────────────────────────────────────────────────
  # Default grid (10..30 kHz step 2, duty 0/50/100 %, dead-band 1 µs):
  python run_power_module_sweep.py --scope-ch 1

  # Single switch, custom grid:
  python run_power_module_sweep.py --switches UTOP \
      --freq-min 10 --freq-step 5 --freq-max 30 \
      --duty-min 20 --duty-step 30 --duty-max 80 --deadband-ns 500

  # Dry-run, no hardware:
  python run_power_module_sweep.py --simulate
"""
import sys
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

import os
import time
import argparse
import datetime
import statistics
from typing import Dict, Any, List

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from hw_protocol import HwTestProtocol
from scope_rs import RohdeScope
from run_gate_scope_check import (
    SWITCHES, DEFAULT_PROBE_ATT, DEFAULT_V_DIV, DEFAULT_OFFSET, DEFAULT_TRIG_LVL,
    PARAM_LIMITS, gen_points, _open_can, _send,
    NAVY, PASS_F, FAIL_F, BORDER, CENTER, _fill,
)

PM_TEST_KEY = "HW_TEST_ALL_POWER_MODULE"   # 0x1B

# The three low-side (bottom) switches run the COMPLEMENT of the commanded
# high-side duty: a "set duty" command sets the high-side (top) on-time, so the
# bottom gate reads (100 - duty) minus the dead-band. Verdicts must account for
# this when the probe is on a bottom switch.
BOTTOM_SWITCHES = {"UBOT", "VBOT", "WBOT"}

# ── standardised display formatting (uniform 2 decimals on every value) ─────
_DECIMALS = 2

def _fmt(key, value):
    """Format a measured value for the report — standardised to 2 decimals."""
    if value != value:                                   # NaN
        return "—"
    if key in ("rise_time", "fall_time"):
        return f"{value * 1e9:.{_DECIMALS}f} ns"
    if key == "frequency":
        return f"{value / 1000.0:.{_DECIMALS}f} kHz"
    if key == "duty":
        return f"{value:.{_DECIMALS}f} %"
    return f"{value:.{_DECIMALS}f} V"                    # voltages


def expected_duty_pct(switch, duty_pct):
    """Expected MEASURED duty (%) for the probed switch: top = commanded,
    bottom = complement (100 - commanded)."""
    return (100.0 - duty_pct) if switch in BOTTOM_SWITCHES else duty_pct


# ── firmware control ───────────────────────────────────────────────────────

def _start_pm_test(bus, protocol, deadband_ns: float,
                   enable_phases: bool = True, startup_delay: float = 1.5):
    """Enter HW_TEST_ALL_POWER_MODULE, enable the phases, pin the dead-band.

    The firmware (test_do_all_power_modules) only drives the gate-driver GPIOs
    when pwm_en_phase_u/v/w_forced >= 1, and on start it forces duty to 0 % and
    freq to the default. So we must explicitly enable the three phases via
    EN_DIS_PHASE (0x09, object index 0 = all phases) and give the supplies /
    drivers time to come up before measuring."""
    flags = (protocol.SF_HW_TEST_PROTOCOL_TEST_SET_ENV_FLAGS_REPORT_ENABLE
             | protocol.SF_HW_TEST_PROTOCOL_TEST_SET_ENV_FLAGS_CLEAR_REBOOT)
    _send(bus, protocol, "SF_HW_TEST_PROTOCOL_RQST_SET_TEST_ENV", 0, flags)
    time.sleep(0.3)
    _send(bus, protocol, "SF_HW_TEST_PROTOCOL_RQST_SET_TEST", 0, PM_TEST_KEY)
    time.sleep(0.3)
    # Enable all three phases (object index 0 → U/V/W); firmware just checks !=0.
    if enable_phases:
        _send(bus, protocol, "SF_HW_TEST_CAN_PROTOCOL_CMD_EN_DIS_PHASE", 0, 1)
        time.sleep(0.3)
    # Dead-band in ns (SET_PWM_DT 0x07). Pinned once; held for the whole grid.
    _send(bus, protocol, "SF_HW_TEST_CAN_PROTOCOL_CMD_SET_PWM_DT", 0, float(deadband_ns))
    time.sleep(0.2)
    # Let the power supplies / gate drivers ramp before the first measurement.
    if startup_delay > 0:
        time.sleep(startup_delay)


def _pin_setpoint(bus, protocol, freq_khz: float, duty_pu: float):
    """Pin one frequency (kHz) + duty (per-unit) setpoint."""
    _send(bus, protocol, "SF_HW_TEST_PROTOCOL_RQST_SET_PWM_FREQ", 0, float(freq_khz))
    time.sleep(0.15)
    _send(bus, protocol, "SF_HW_TEST_PROTOCOL_RQST_SET_PWM_DUTY", 0, float(duty_pu))
    time.sleep(0.15)


def _stop_pm_test(bus, protocol):
    _send(bus, protocol, "SF_HW_TEST_PROTOCOL_RQST_SET_TEST", 0, "HW_TEST_NO_TEST")
    time.sleep(0.1)


# ── per-setpoint checking ──────────────────────────────────────────────────

def _check(key, value, setpoints):
    """(passed, note). `setpoints` holds the commanded values to verify
    (e.g. {"frequency": 16000.0, "duty": 50.0}); everything else uses
    PARAM_LIMITS. NaN always fails (no reading)."""
    if value != value:                       # NaN
        return False, "no reading"
    if key in setpoints and setpoints[key] is not None:
        sp = setpoints[key]
        if key == "duty":
            ok = abs(value - sp) <= 5.0      # ±5 percentage-points
            return ok, f"setpoint {sp:g}% ±5pp"
        tol = 0.05 * sp
        ok = abs(value - sp) <= tol
        return ok, f"setpoint {sp:g} ±5%"
    lo, hi, _u = PARAM_LIMITS.get(key, (None, None, ""))
    if lo is not None and value < lo:
        return False, f"< {lo}"
    if hi is not None and value > hi:
        return False, f"> {hi}"
    return True, "ok"


def _eval_checks(switch, freq_khz, duty_pct, meas):
    """Switch-aware per-parameter checks + overall verdict for one record.

    The duty setpoint is the EXPECTED measured duty for the probed switch
    (top = commanded, bottom = 100 - commanded); the ±5 pp tolerance in
    _check() absorbs the dead-band offset. At 0 %/100 % duty there are no
    switching edges, so frequency/duty are report-only."""
    degenerate = duty_pct <= 0.0 or duty_pct >= 100.0
    setpoints = {} if degenerate else {
        "frequency": freq_khz * 1000.0,
        "duty": expected_duty_pct(switch, duty_pct)}
    checks = {k: _check(k, v, setpoints) for k, v in meas.items()}
    if degenerate:
        checks["frequency"] = (True, "n/a (DC)")
        checks["duty"] = (True, "n/a (DC)")
    verdict = all(ok for ok, _ in checks.values())
    return checks, verdict


def recompute_verdict(rec):
    """Re-evaluate a stored record switch-aware (in place) and return it.
    Used when regenerating a report from previously-captured session data."""
    checks, verdict = _eval_checks(rec.get("switch", ""), rec["freq_khz"],
                                   rec["duty_pct"], rec["meas"])
    rec["checks"] = checks
    rec["verdict"] = verdict
    return rec


def _tdiv_for(freq_khz):
    f_hz = freq_khz * 1000.0
    return max((1.0 / f_hz) * 3 / 10.0, 1e-7) if f_hz else 20e-6


def build_grid(freq_points, duty_points):
    """Cartesian product of frequency (kHz) × duty (%) setpoints."""
    return [(f, d) for f in freq_points for d in duty_points]


def measure_pm_setpoint(scope, bus, protocol, scope_ch,
                        freq_khz, duty_pct, deadband_ns,
                        simulate, edge_t_div=None, settle_s=0.8,
                        clean_restart=False, switch=None):
    """Pin one (freq, duty) setpoint of the power-module test, measure the
    scope, and return a result record.

    Duty is given in PERCENT and sent PER-UNIT (÷100). At 0 %/100 % duty there
    are no switching edges, so frequency/duty are report-only (the verdict
    ignores them).

    clean_restart : if True, fully stop+restart the test (and re-enable the
        phases via _start_pm_test) before pinning this setpoint. This makes
        every point start from the firmware's clean state — the only reliable
        way to step the switching frequency, because changing freq live during
        a running ALL_POWER_MODULE test latches the gate (the duty compare is
        not rescaled to the new period). If False, the test must already be
        running (set once by _start_pm_test)."""
    duty_pu = duty_pct / 100.0

    if not simulate:
        if clean_restart:
            _stop_pm_test(bus, protocol)
            time.sleep(0.2)
            _start_pm_test(bus, protocol, deadband_ns)   # SET_TEST + enable + ramp
        _pin_setpoint(bus, protocol, freq_khz, duty_pu)
    scope.configure_timebase(_tdiv_for(freq_khz))
    time.sleep(settle_s)
    meas = scope.measure(scope_ch, edge_t_div=edge_t_div)

    # Switch-aware verdict (top = commanded duty, bottom = complement).
    checks, verdict = _eval_checks(switch, freq_khz, duty_pct, meas)

    return dict(freq_khz=freq_khz, duty_pct=duty_pct, deadband_ns=deadband_ns,
                meas=meas, checks=checks, verdict=verdict, switch=switch)


# ── report ─────────────────────────────────────────────────────────────────

def write_report(results: List[Dict[str, Any]], meta: Dict[str, Any],
                 out_path: str):
    # ── one font family (Arial) and consistent sizes on EVERY cell ──────────
    FONT = "Arial"
    F_TITLE   = Font(name=FONT, size=14, bold=True, color=NAVY)
    F_META_L  = Font(name=FONT, size=10, bold=True)
    F_META_V  = Font(name=FONT, size=10)
    F_HDR     = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    F_DATA    = Font(name=FONT, size=10)
    F_SWITCH  = Font(name=FONT, size=10, bold=True)
    F_FAIL    = Font(name=FONT, size=10, bold=True, color="9C0006")
    F_PASS_V  = Font(name=FONT, size=10, bold=True, color="006100")

    wb = Workbook()
    ws = wb.active
    ws.title = "Power Module Sweep"

    tc = ws.cell(row=1, column=1,
                 value="Inverter Gen3 — Power-module sweep "
                       "(HW_TEST_ALL_POWER_MODULE 0x1B)")
    tc.font = F_TITLE
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=12)

    r = 3
    for k, v in [("Unit S/N", meta["unit_sn"]), ("Operator", meta["operator"]),
                 ("Scope", meta["scope_idn"]), ("Probe atten", meta["probe_atten"]),
                 ("Dead-band", meta["deadband"]), ("Sweep", meta["sweep_str"]),
                 ("Date", meta["date"])]:
        lc = ws.cell(row=r, column=1, value=k)
        lc.font = F_META_L; lc.fill = _fill("D9E1F2"); lc.border = BORDER
        c = ws.cell(row=r, column=2, value=str(v))
        c.border = BORDER; c.font = F_META_V
        ws.merge_cells(start_row=r, end_row=r, start_column=2, end_column=12)
        r += 1

    r += 1
    cols = ["Switch", "Freq cmd", "Duty cmd", "DB (ns)", "Frequency", "Duty",
            "Rise", "Fall", "V-high", "V-low", "V-pp", "Verdict"]
    widths = [9, 10, 10, 8, 13, 10, 10, 10, 10, 10, 10, 10]
    for ci, (lbl, w) in enumerate(zip(cols, widths), 1):
        c = ws.cell(row=r, column=ci, value=lbl)
        c.font = F_HDR; c.fill = _fill(NAVY); c.alignment = CENTER; c.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    hdr_row = r

    keys = ["frequency", "duty", "rise_time", "fall_time", "v_high", "v_low", "v_pp"]
    MEAS_FIRST = 5   # measurement cols start at spreadsheet column 5
    for rec in results:
        r += 1
        verdict_ok = rec["verdict"]
        rowfill = _fill(PASS_F if verdict_ok else FAIL_F)
        # leading command columns — standardised decimals
        for ci, val in enumerate([rec["switch"],
                                  f"{rec['freq_khz']:.{_DECIMALS}f} kHz",
                                  f"{rec['duty_pct']:.{_DECIMALS}f} %",
                                  f"{rec['deadband_ns']:.0f}"], 1):
            cc = ws.cell(row=r, column=ci, value=val)
            cc.fill = rowfill; cc.border = BORDER; cc.alignment = CENTER
            cc.font = F_SWITCH if ci == 1 else F_DATA
        for i, key in enumerate(keys):
            ci = MEAS_FIRST + i
            val = rec["meas"].get(key, float("nan"))
            cell = ws.cell(row=r, column=ci, value=_fmt(key, val))
            ok = rec["checks"].get(key, (True, ""))[0]
            cell.fill = _fill(PASS_F if ok else FAIL_F)
            cell.border = BORDER; cell.alignment = CENTER
            cell.font = F_DATA if ok else F_FAIL
        vc = ws.cell(row=r, column=12, value=("PASS" if verdict_ok else "FAIL"))
        vc.fill = rowfill; vc.border = BORDER; vc.alignment = CENTER
        vc.font = F_PASS_V if verdict_ok else F_FAIL
    ws.freeze_panes = ws.cell(row=hdr_row + 1, column=1)

    # ── summary statistics rows (mean / abs std / relative std) ─────────────
    # Per measured column, over all points (NaN ignored). Note: frequency and
    # duty are swept by design, so their spread reflects the sweep range; the
    # rise/fall/voltage columns characterise gate consistency.
    F_STAT_L = Font(name=FONT, size=10, bold=True)
    F_STAT_V = Font(name=FONT, size=10, bold=True, color=NAVY)
    SUM_FILL = _fill("D9E1F2")
    series = {k: [rec["meas"].get(k, float("nan")) for rec in results] for k in keys}
    series = {k: [v for v in vals if v == v] for k, vals in series.items()}   # drop NaN

    def _stat(key, kind):
        vals = series.get(key, [])
        if not vals:
            return "—"
        mean = sum(vals) / len(vals)
        if kind == "mean":
            return _fmt(key, mean)
        sd = statistics.pstdev(vals) if len(vals) > 1 else 0.0
        if kind == "abs":
            return _fmt(key, sd)
        rel = (sd / abs(mean) * 100.0) if mean else float("nan")
        return "—" if rel != rel else f"{rel:.{_DECIMALS}f} %"

    # Frequency and duty are swept setpoints — their mean/std is not meaningful,
    # so leave them blank in the summary rows; stats only on the gate-quality
    # columns (rise, fall, V-high, V-low, V-pp).
    STATS_EXCLUDE = {"frequency", "duty"}
    for label, kind in [("Mean", "mean"), ("Std (abs)", "abs"), ("Std (rel)", "rel")]:
        r += 1
        lc = ws.cell(row=r, column=1, value=label)
        lc.fill = SUM_FILL; lc.border = BORDER; lc.alignment = CENTER; lc.font = F_STAT_L
        for ci in (2, 3, 4, 12):                      # cmd cols + verdict: blank
            bc = ws.cell(row=r, column=ci, value="")
            bc.fill = SUM_FILL; bc.border = BORDER
        for i, key in enumerate(keys):
            val = "" if key in STATS_EXCLUDE else _stat(key, kind)
            cell = ws.cell(row=r, column=MEAS_FIRST + i, value=val)
            cell.fill = SUM_FILL; cell.border = BORDER
            cell.alignment = CENTER; cell.font = F_STAT_V

    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
    wb.save(out_path)
    return os.path.abspath(out_path)


# ── main ─────────────────────────────────────────────────────────────────

def main():
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    # Frequency sweep (kHz)
    p.add_argument("--freq-min", type=float, default=10.0, help="kHz (default 10)")
    p.add_argument("--freq-step", type=float, default=10.0, help="kHz (default 10 → 10,20,30)")
    p.add_argument("--freq-max", type=float, default=30.0, help="kHz (default 30)")
    p.add_argument("--freq-points", nargs="+", type=float, default=None,
                   help="Explicit frequency points (kHz); overrides min/step/max")
    # Duty sweep (PERCENT on CLI, sent per-unit to firmware)
    p.add_argument("--duty-min", type=float, default=25.0, help="%% (default 25)")
    p.add_argument("--duty-step", type=float, default=25.0, help="%% (default 25)")
    p.add_argument("--duty-max", type=float, default=75.0, help="%% (default 75)")
    p.add_argument("--duty-points", nargs="+", type=float, default=None,
                   help="Explicit duty points (%%); overrides min/step/max")
    # Dead-band (ns)
    p.add_argument("--deadband-ns", type=float, default=1000.0,
                   help="Dead-band in ns (default 1000 = 1 µs)")
    # Scope
    p.add_argument("--edge-tdiv", type=float, default=100e-9,
                   help="Fast timebase (s/div) for accurate rise/fall "
                        "(default 100e-9). 0 disables edge-zoom.")
    p.add_argument("--scope-ch", type=int, default=1)
    p.add_argument("--resource", default=None, help="Explicit VISA resource string")
    p.add_argument("--backend", default="@ivi",
                   help="VISA backend: '@ivi' (vendor VISA — required for USB "
                        "on this bench, default) or '@py' (pyvisa-py)")
    p.add_argument("--probe-atten", type=float, default=DEFAULT_PROBE_ATT)
    p.add_argument("--v-div", type=float, default=DEFAULT_V_DIV)
    p.add_argument("--offset", type=float, default=DEFAULT_OFFSET)
    p.add_argument("--trig-level", type=float, default=DEFAULT_TRIG_LVL)
    # Misc
    p.add_argument("--unit-sn", default="PCB-B1")
    p.add_argument("--operator", default="Carlos Miguel Espinar")
    p.add_argument("--switches", nargs="+", default=SWITCHES,
                   help="Subset/order of switches to check")
    p.add_argument("--no-prompts", action="store_true",
                   help="Skip the per-switch 'move the probe' prompt")
    p.add_argument("--simulate", action="store_true",
                   help="No hardware: canned scope data, CAN skipped")
    args = p.parse_args()

    # Build the sweep grid (explicit points override the min/step/max sweep).
    freq_points = (args.freq_points if args.freq_points is not None
                   else gen_points(args.freq_min, args.freq_max, args.freq_step))
    duty_points = (args.duty_points if args.duty_points is not None
                   else gen_points(args.duty_min, args.duty_max, args.duty_step))
    edge_t_div = args.edge_tdiv if args.edge_tdiv and args.edge_tdiv > 0 else None

    grid = build_grid(freq_points, duty_points)
    sweep_str = (f"freq {freq_points} kHz × duty {duty_points} %  "
                 f"({len(freq_points)}×{len(duty_points)} = {len(grid)} pts/switch)")

    protocol = HwTestProtocol()
    bus = None if args.simulate else _open_can()
    scope = RohdeScope(resource=args.resource, backend=args.backend,
                       simulate=args.simulate).open()

    print("=" * 72)
    print("  Power-module sweep — HW_TEST_ALL_POWER_MODULE (0x1B)")
    print(f"  Scope     : {scope.idn()}")
    print(f"  Scope ch  : CH{args.scope_ch}   probe {args.probe_atten:g}:1   "
          f"{args.v_div} V/div, offset {args.offset} V")
    print(f"  Dead-band : {args.deadband_ns:g} ns")
    print(f"  Sweep     : {sweep_str}")
    print(f"  Switches  : {', '.join(args.switches)}")
    print("=" * 72)

    # Full scope pre-config (reset + vertical + trigger + 7 on-screen
    # measurement slots). Timebase is then adjusted per setpoint below.
    scope.full_setup(args.scope_ch, v_div=args.v_div, offset=args.offset,
                     probe_atten=args.probe_atten, t_div=_tdiv_for(freq_points[0]),
                     trig_level=args.trig_level)

    if not args.simulate:
        _start_pm_test(bus, protocol, args.deadband_ns)

    results = []
    try:
        for sw in args.switches:
            print()
            print("  " + "═" * 68)
            print(f"  ⚠  Move the voltage probe to the  {sw}  gate (isolated PWM).")
            print("  " + "═" * 68)
            if not args.simulate and not args.no_prompts:
                try:
                    input(f"  Press ENTER when the probe is on {sw} … ")
                except (EOFError, KeyboardInterrupt):
                    print("\n  Aborted by operator."); break

            for freq_khz, duty_pct in grid:
                rec = measure_pm_setpoint(scope, bus, protocol, args.scope_ch,
                                          freq_khz, duty_pct, args.deadband_ns,
                                          args.simulate, edge_t_div=edge_t_div)
                rec["switch"] = sw
                results.append(rec)
                meas = rec["meas"]
                print(f"    [f={freq_khz:>5g} kHz  duty={duty_pct:>5g} %]  "
                      f"f={_fmt('frequency', meas['frequency'])}  "
                      f"duty={_fmt('duty', meas['duty'])}  "
                      f"Vhi={_fmt('v_high', meas['v_high'])}  "
                      f"Vlo={_fmt('v_low', meas['v_low'])}  "
                      f"rise={_fmt('rise_time', meas['rise_time'])}  "
                      f"→ {'PASS' if rec['verdict'] else 'FAIL'}")
    finally:
        if not args.simulate:
            _stop_pm_test(bus, protocol)
            bus.close()
        scope.close()

    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M")
    sn_safe = args.unit_sn.replace("/", "-").replace(" ", "_")
    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "results",
                       f"PowerModuleSweep_{sn_safe}_{ts}.xlsx")
    meta = dict(unit_sn=args.unit_sn, operator=args.operator,
                scope_idn=scope.idn(), probe_atten=f"{args.probe_atten:g}:1",
                deadband=f"{args.deadband_ns:g} ns ({args.deadband_ns/1000:g} µs)",
                sweep_str=sweep_str,
                date=datetime.datetime.now().isoformat(timespec="seconds"))
    out = write_report(results, meta, out)

    n_pass = sum(1 for r in results if r["verdict"])
    print()
    print("=" * 72)
    print(f"  POWER-MODULE SWEEP COMPLETE — {n_pass}/{len(results)} setpoints PASS")
    print(f"  Report: {out}")
    print("=" * 72)
    return 0 if n_pass == len(results) else 1


if __name__ == "__main__":
    sys.exit(main())
